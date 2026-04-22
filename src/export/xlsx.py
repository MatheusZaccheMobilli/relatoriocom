"""Geração de XLSX do relatório de comissão — mesma identidade visual do PDF."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import RelatorioData


# Paleta Mobílli (hex)
LARANJA = "FF6600"
LARANJA_ESCURO = "CC5200"
CINZA_CLARO = "F0F0F0"
VERMELHO_SUAVE = "FFCCCC"
BRANCO = "FFFFFF"
PRETO = "1E1E1E"

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def _brl(valor: Decimal) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_data(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else "-"


def _mes_ano(d: date) -> str:
    return f"{MESES_PT[d.month]}/{d.year}"


def _nivel_label(nome: str) -> str:
    return {
        "Ouro": "🥇 Ouro",
        "Prata": "🥈 Prata",
        "Bronze": "🥉 Bronze",
    }.get(nome, nome)


def gerar_xlsx(relatorio: RelatorioData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comissão"

    thin = Side(border_style="thin", color="BBBBBB")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor=LARANJA)
    fill_header_escuro = PatternFill("solid", fgColor=LARANJA_ESCURO)
    fill_cinza = PatternFill("solid", fgColor=CINZA_CLARO)
    fill_devolvido = PatternFill("solid", fgColor=VERMELHO_SUAVE)

    # ── Título
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "Relatório de Comissão de Vendedores — Mobílli"
    c.font = Font(name="Calibri", size=16, bold=True, color=BRANCO)
    c.fill = fill_header_escuro
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = f"Pagamento em {_mes_ano(relatorio.competencia)}"
    c.font = Font(name="Calibri", size=11, italic=True)
    c.alignment = Alignment(horizontal="center")

    # ── Dados do vendedor + meta (blocos)
    row = 4
    ws.cell(row=row, column=1, value="Vendedor").font = Font(bold=True)
    ws.cell(row=row, column=2, value=relatorio.vendedor.nome)
    ws.cell(row=row, column=4, value="Competência").font = Font(bold=True)
    ws.cell(row=row, column=5, value=_mes_ano(relatorio.competencia))

    if relatorio.vendedor.cpf:
        row += 1
        ws.cell(row=row, column=1, value="CPF").font = Font(bold=True)
        ws.cell(row=row, column=2, value=relatorio.vendedor.cpf)

    row += 2
    ws.cell(row=row, column=1, value="Meta").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{relatorio.nivel.qtd_meta} captações")
    ws.cell(row=row, column=4, value="Total geral").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{relatorio.nivel.qtd_atingida} captações")
    ws.cell(row=row, column=7, value="% Atingido").font = Font(bold=True)
    ws.cell(row=row, column=8, value=f"{relatorio.nivel.percentual_atingido}%")
    ws.cell(row=row, column=10, value="Nível").font = Font(bold=True)
    ws.cell(row=row, column=11, value=_nivel_label(relatorio.nivel.nome))

    row += 2
    ws.cell(row=row, column=1, value="Itens para comissão").font = Font(bold=True)
    ws.cell(row=row, column=2, value=relatorio.negocios_fechados)
    ws.cell(row=row, column=4, value="Devolvidos").font = Font(bold=True)
    ws.cell(row=row, column=5, value=relatorio.negocios_encerrados)

    # ── Tabela de itens
    row += 2
    headers = [
        "Tipo", "Parcela", "Plano", "Nome do Cliente", "Placa",
        "Data Locação", "Data Devolução",
        "Valor Base", "Parcelas", "Comissão", "Status",
    ]
    header_row = row
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=BRANCO)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[row].height = 22

    row += 1
    for item in relatorio.itens:
        if item.tipo_operacao == "Locação":
            plano = "Semanal" if item.plano_semanal else "Mensal"
            parcelas = item.qtd_parcelas_pagas
        else:
            plano = "—"
            parcelas = "—"

        status = "DEVOLVIDO" if item.devolvido else "Ativo"

        valores = [
            item.tipo_operacao,
            item.parcela,
            plano,
            item.nome_cliente,
            item.placa,
            _fmt_data(item.data_locacao),
            _fmt_data(item.data_devolucao),
            float(item.valor_base),
            parcelas,
            float(item.valor_comissao),
            status,
        ]
        for col, v in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border_thin
            if item.devolvido:
                cell.fill = fill_devolvido
        ws.cell(row=row, column=8).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=10).number_format = 'R$ #,##0.00'
        # alinhamentos
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=11).alignment = Alignment(horizontal="center")
        row += 1

    # ── Total
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value="TOTAL A RECEBER").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).fill = fill_cinza

    total_cell = ws.cell(row=row, column=10, value=float(relatorio.total_comissao))
    total_cell.font = Font(bold=True, size=12, color=LARANJA_ESCURO)
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.alignment = Alignment(horizontal="center")
    total_cell.fill = fill_cinza
    ws.cell(row=row, column=11).fill = fill_cinza

    # ── Larguras das colunas
    larguras = {
        1: 12,   # Tipo
        2: 9,    # Parcela
        3: 10,   # Plano
        4: 42,   # Cliente
        5: 11,   # Placa
        6: 13,   # Data Locação
        7: 14,   # Data Devolução
        8: 13,   # Valor Base
        9: 10,   # Parcelas
        10: 13,  # Comissão
        11: 12,  # Status
    }
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Exportar
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
